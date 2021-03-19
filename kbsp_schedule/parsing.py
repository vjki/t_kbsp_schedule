#!/usr/bin/env/ python
# -*- coding: utf-8 -*-
#
# # # # # # # # # # # # # # # # # # # # # # # # # # # # #
# Source: https://github.com/skvozsneg/t_kbsp_schedule  #
# # # # # # # # # # # # # # # # # # # # # # # # # # # # #
#
#
import json
import openpyxl

from os import path, listdir

# TODO: FILE Needs refactoring

# --- Functions ---
def write_in_json(json_dir: str, schedule_d: dict, course: int):
    """Write groups schedule in json file

    • json_dir      - string which contain way to json dir

    • schedule_d    - dictionary with groups schedule data

    • course        - group number of the course
    """
    try:
        json_dir_full = path.join(json_dir, str(course))
        json_name = schedule_d['Group'] + '.json'
        with open(path.join(json_dir_full, json_name), 'w', encoding='utf-8') as f:
            json.dump(schedule_d, f, ensure_ascii=False, indent=4)
        return True
    except:
        return False


def pars_for_cells(schedule_dir: str):
    """GENERATOR. Gets cells with group name.
    Go throw th 2-nd row in xlsx and catch cells with group names. Yield
    dictionary (d) with file name and cells.
        { 'full_path': full_path, 'course': sub_dir, 'cells': [] }

    • schedule_dir - string which contain way to schedule dir

    """
    for sub_dir in range(1, 6):
        current_dir = path.join(schedule_dir, str(sub_dir))
        for file_name in listdir(current_dir):
            full_path = path.join(current_dir, file_name)
            d = {'full_path': full_path, 'course': sub_dir, 'cells': []}
            wb = openpyxl.load_workbook(full_path, read_only=True)
            sheet = wb.active
            for row in sheet.iter_rows(2):
                for cell in row:
                    if cell.value is not None and '-' in str(cell.value):
                        d['cells'].append(cell.coordinate)
                break
            yield d
            wb.close()


def pars_main(d_cells: dict, json_dir: str) -> bool:
    """Parsing Excel.
    Function collects data from .xlsx file and return it as a list of dictionary.

    • d_cells   - dictionary with full path to file and its cells with group name

    • json_dir  - path to json storage.

    Return:

    • d         - dictionary with group schedule data's. Every list there collect four type of information.
                    1: Subject name;
                    2: Subject type;
                    3: Professor name;
                    4: Audience;
                        view: {'Group'(group name): str,
                                11(monday):{
                                    1(first lesson): {
                                        111(odd week): list,
                                        112(even week): list
                                    },
                                    2(second lesson): {
                                        111(odd week): list,
                                        112(even week): list
                                    },
                                            ...
                                    6(sixth lesson): {
                                        111(odd week): list,
                                        112(even week): list
                                    },
                                }
                                12(tuesday):{...}
                                13(wednesday):{...}
                                        ...
                                16(saturday):{...}
                        }
    """
    full_path = d_cells['full_path']
    cell_coordinate = d_cells['cells']
    course = d_cells['course']
    wb = openpyxl.load_workbook(full_path, read_only=True)
    sheet = wb.active
    for el in cell_coordinate:
        d = {x: {x: {}.fromkeys(range(111, 113), []) for x in range(1, 7)} for x in range(11, 17)}
        d.update({'Group': None})
        count_nmb_weeks, count_weeks, count_lessons = 111, 11, 1
        d['Group'] = sheet[el].value
        try:
            xy = openpyxl.utils.coordinate_to_tuple(el)
            # Выделение "зоны" группы в екселе, в которой содежится вся инфа и считывание."
            up = openpyxl.utils.get_column_letter(xy[1]) + str(xy[0] + 2)           # верхняя левая координата
            down = openpyxl.utils.get_column_letter(xy[1] + 3) + str(xy[0] + 73)    # нижняя правая координата
            for row in sheet[up : down]:
                for cell in row:
                    val = cell.value
                    # TODO: ЗАПИСЫВАЕТ ОДНОВРЕМЕННО НА ЧЕТНУЮ И НЕЧЕТНУЮ НЕДЕЛИ. ИСПРАВИТЬ!!!
                    d[count_weeks][count_lessons][count_nmb_weeks].append(val if val is not None else '-')
                    # if cell.value is None:
                    #     d[count_weeks][count_lessons][count_nmb_weeks].append('-')
                    # else:
                    #     d[count_weeks][count_lessons][count_nmb_weeks].append(cell.value)
                if count_nmb_weeks == 111:
                    count_nmb_weeks += 1
                else:
                    count_nmb_weeks -= 1
                    count_lessons += 1
                if count_lessons > 6:
                    count_weeks += 1
                    count_lessons = 1
            write_in_json(json_dir, d, course)
        except:
            wb.close()
            return False
    wb.close()

            
